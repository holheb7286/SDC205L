from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
import pandas as pd
import matplotlib.pyplot as plt

# This function writes comma-separated data to a CSV file.
def insertData(path, data):
    try:
        with open(path, 'a') as file:
            file.write(data + "\n")
        return True
    except:
        print("Error: Unable to write to file.")
        return False

# This function reads and displays the contents of a CSV file.
def viewData(path):
    try:
        print(f"Reading from file: {path}")
        with open(path, 'r') as file:
            for line in file:
                print(line.strip())
    except:
        print("Error: Unable to read file.")

# This function converts Fahrenheit to Celsius.
def convertData(data):
    converted_value = (data - 32) * 5 / 9
    return converted_value

# This function gets user input, converts the value, and saves the data to a CSV file.
def getInput():
    path = "C:\\PythonFiles\\ZooData.csv"
    entries = int(input("How many entries are being entered? "))

    for i in range(entries):
        entry_date = input("Enter the date: ")
        value = float(input("Enter the temperature in Fahrenheit: "))

        # convertData(data) requires a numerical value and returns the converted Celsius value.
        converted_value = convertData(value)

        data = f"{entry_date},{value},{converted_value:.2f}"

        try:
            if insertData(path, data):
                print(f"The following data was saved at {datetime.now()}: {data}.")
        except:
            print("Error: Unable to save data.")

# This function asks the user for a graph type and generates a report chart from the CSV file.
# Arguments: path (string) is the path to the CSV file.
# Return value: none.
def generateReport(path):
    chart_type = input("Enter chart type (line or bar): ")
    createChart(path, chart_type)

# This function creates an Excel spreadsheet and chart from CSV data.
# Arguments: path (string) is the path to the CSV file, chart_type (string) is either "line" or "bar".
# Return value: none.
def createChart(path, chart_type):
    try:
        data_source = input("Choose data source: 1 for Fahrenheit, 2 for Celsius: ")

        dates = []
        values = []

        with open(path, 'r') as file:
            for line in file:
                row = line.strip().split(',')

                if len(row) == 3:
                    dates.append(row[0])

                    if data_source == "1":
                        values.append(float(row[1]))
                    elif data_source == "2":
                        values.append(float(row[2]))
                    else:
                        print("Error: Invalid data source selected.")
                        return

        wb = Workbook()
        ws = wb.active
        ws.title = "ChartData"

        ws["A1"] = "Date"

        if data_source == "1":
            ws["B1"] = "Fahrenheit"
            y_axis_title = "Temperature (F)"
        else:
            ws["B1"] = "Celsius"
            y_axis_title = "Temperature (C)"

        for i in range(len(dates)):
            ws.cell(row=i + 2, column=1, value=dates[i])
            ws.cell(row=i + 2, column=2, value=values[i])

        data = Reference(ws, min_col=2, min_row=2, max_row=len(values) + 1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(dates) + 1)

        if chart_type.lower() == "bar":
            chart = BarChart()
        elif chart_type.lower() == "line":
            chart = LineChart()
        else:
            print("Error: Invalid chart type selected.")
            return

        chart.add_data(data)
        chart.set_categories(labels)

        chart.title = f"{student_id} {datetime.now().strftime('%m/%d/%Y')}"
        chart.x_axis.title = "Date"
        chart.y_axis.title = y_axis_title

        ws.add_chart(chart, "D2")
        wb.save("C:\\PythonFiles\\final.xlsx")

        print("Chart created successfully in final.xlsx")
    except:
        print("Error: Unable to create chart.")

# Display application name
student_id = input("Enter your Student ID: ")
print(f"{student_id}'s Spreadsheet Automation Menu")

# Store menu options
menu_options = [
    "1. Input Data",
    "2. View Current Data",
    "3. Generate Report"
]

# Print menu options via loop
for item in menu_options:
    print(item)

# Retrieve menu selection
option = int(input("Select a menu option (1-3): "))

# Get current date and time
current_time = str(datetime.now())

# Previous error-checking code
if option == 1:
    print(f"You selected menu option 1 at {current_time}")
elif option == 2:
    print(f"You selected menu option 2 at {current_time}")
elif option == 3:
    print(f"You selected menu option 3 at {current_time}")
else:
    print("Error: Invalid choice selected.")

# Menu flow control
path = "C:\\PythonFiles\\ZooData.csv"

if option == 1:
    getInput()
elif option == 2:
    viewData(path)
elif option == 3:
    generateReport(path)
else:
    print("Error: The chosen functionality is not implemented yet")
